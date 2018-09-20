from PlatformApp.mainApp import app, db, update_ret, add_stock, update_score, create_stock_info
from PlatformApp.models import User, Tickers, Role, Transactions, Stock, Vote, RecentVote, AnalystFile, FundFile, ticker_identifier, roles_users
from flask_script import Manager, prompt_bool
from openpyxl import load_workbook, Workbook
import requests
from spreadsheet import sheet
from dues import transactions, member_info
from votingchallenge import get_users_from_vote, get_ticker_identifier, get_roles_users
import httplib2
from boxsdk import OAuth2
from boxsdk import Client
from sqlalchemy import create_engine, Column, Integer, String, Boolean, Float, insert
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy import func
from lxml import html
from pandas_datareader import data
import pandas as pd

# from votingchallenge import votesession, get_users_from_vote
# from signin import signsession

manager = Manager(app)

# platforme = create_engine('postgresql+psycopg2://qcksqjzmfkdxdo:password@127.0.0.1:5432/myusit')
# Session = sessionmaker(bind=platforme)
# platformsession = Session()

@manager.command
def initdb():
    db.create_all()
    admin = Role(name="Admin", description="Gets to look at all the rankings.")
    #db.session.add(User(firstName="Arnav", lastName="Jain",email="arnav_jain@utexas.edu", eid="aj24872", password="test11", roles=[admin], attendance = 0, dues = 0, atLatestMeeting = False, rowOnSheet = 0, analyst = 'Srija Nalla', fund='TMT'))
    db.session.add(User(firstName="Catherine", lastName="Cheng",email="president@usiteam.org", eid="1111111", password="test11", roles=[admin], attendance = 0, dues = 0, atLatestMeeting = False, rowOnSheet = 0, analyst = 'Srija Nalla', fund='TMT'))
    db.session.add(User(firstName="Eric", lastName="Sun",email="vicepresident@usiteam.org", eid="2222222", password="test11", roles=[admin], attendance = 0, dues = 0, atLatestMeeting = False, rowOnSheet = 0, analyst = 'Srija Nalla', fund='TMT'))
    db.session.add(User(firstName="Matt", lastName="Hopp",email="investment@usiteam.org", eid="3333333", password="test11", roles=[admin], attendance = 0, dues = 0, atLatestMeeting = False, rowOnSheet = 0, analyst = 'Srija Nalla', fund='TMT'))
    db.session.add(User(firstName="Jessica", lastName="Huang",email="technology@usiteam.org", eid="4444444", password="test11", roles=[admin], attendance = 0, dues = 0, atLatestMeeting = False, rowOnSheet = 0, analyst = 'Srija Nalla', fund='TMT'))
    db.session.commit()
    refreshdb()
    print('Initialized the database')

@manager.command
def openfile():
    try: 
        name = str(raw_input("File path: "))
        wb = load_workbook(filename = name)
        ws = wb.active
        print('file ' + name + ' uploaded ')
    except:
        print('file not uploaded')

def get_info_server(ticker):

    info = {}

    if db.session.query(Stock).filter_by(ticker = ticker).count() > 0:
        found_stock = db.session.query(Stock).filter_by(ticker = ticker).first()
        info['name'] = found_stock.name
        info['price'] = found_stock.price
        info['datetime'] = found_stock.datetime
        info['gain'] = found_stock.change
        info['percentchange'] = found_stock.percentChange
    else:
        info['name'] = "N/A"
        info['price'] = 0.00
        info['datetime'] = "N/A"
        info['gain'] = 0.00
        info['percentchange'] = 0.00

    return info    


@manager.command
def get_price():
    ticker = str(raw_input("Ticker: "))
    
    else:
        price = 0
        print('price not found')
    #return price

@manager.command
def get_info():
    ticker = str(raw_input("Ticker: "))
    # rjson = get_json(ticker)
    info = {}

    # s = Stock(ticker)
    urlStock = "http://www.nasdaq.com/symbol/{}".format(ticker)
    page = requests.get(urlStock)
    tree = html.fromstring(page.content)

    try:
        price = round(float(tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0].split("$")[1]), 2)
        print(price)
        change = round(float(tree.xpath('//div[@id="qwidget_netchange"]/text()')[0]), 2)
        pchange = round(float(tree.xpath('//div[@id="qwidget_percent"]/text()')[0].split("%")[0]), 2)
        change_class = tree.xpath('//div[@id="qwidget_netchange"]')[0].get('class')

        print(str(price), str(change), str(pchange))

        negative = True

        if change_class == 'qwidget-cents qwidget-Red':
            negative = True
        elif change_class == 'qwidget-cents qwidget-Green':
            negative = False
        else:
            print("I guess it does not exist...")

        # 0: Get name
        url = "http://d.yimg.com/autoc.finance.yahoo.com/autoc?query={}&region=1&lang=en".format(ticker)
        print("Ticker: ", ticker)
        result = requests.get(url).json()

        for x in result['ResultSet']['Result']:
            if x['symbol'] == ticker:
                info['name'] = truncate(x['name'])

        if info['name'] == "" or info['name'] == None:
            raise ValueError('Did not obtain a real value!')


        # 1: Get price
        # info['price'] = float(rjson[0][u'l'])
        # info['price'] = get_price(ticker)
        info['price'] = price

        if info['price'] == 0 or info['price'] == None:
            raise ValueError('Did not obtain a real value!')

        # 2: Get datetime
        # info['datetime'] = rjson[0][u'lt']
        info['datetime'] = getdatetime(ticker)

        if info['datetime'] == "" or info['datetime'] == None:
            raise ValueError('Did not obtain a real value!')

        # 3: Get gain
        # change = rjson[0][u'c']
        # if change is None:
        #     info['gain'] = 0
        # c = change.split("+")
        # if (len(c) > 1):
        #     info['gain'] = float(c[1])
        # info['gain'] = float(change)

        if negative:
            change = -1 * change
        
        if change is None:
            info['gain'] = 0
        else:
            info['gain'] = change

        if info['gain'] == None:
            raise ValueError('Did not obtain a real value!')

            # 4: Get percent change
            # info['percentchange'] = float(rjson[0][u'cp'])
            # try:
            #     percentChange = stock.get_percent_change()
            #     percentChange = percentChange.split("%")[0]
            #     if len(percentChange.split("+")) > 1:
            #         percentChange = percentChange.split("+")[1]
            #     elif len(percentChange.split("-")) > 1:
            #         percentChange = percentChange.split("-")[1]

            #     info['percentchange'] = float(percentChange)
            # except:
            #     info['percentchange'] = stock.get_percent_change()
        
        if negative:
            info['percentchange'] = -pchange
        else:
            info['percentchange'] = pchange

        if info['percentchange'] == None:
            raise ValueError('Did not obtain a real value!')
    except:
        print("WENT TO THE EXCEPTION!")
        info = get_info_server(ticker)

    return info

@manager.command
def get_ag_materials():
    dev_token = str(raw_input("Developer token: "))

    oauth = OAuth2(
      client_id='l91tu18y9v1t9yth4w3xub87jbyln18k',
      client_secret='BYSJR7wr4Tl7Socbw3l87FYK01OOE91r',
      access_token=dev_token
    )

    client = Client(oauth)
    root_folder = client.folder(folder_id='0')
    USIT_folder = root_folder.get_items(limit=100, offset=0)[0]
    # for item in USIT_folder.get_items(limit=100, offset=0):
    #     print item.get()['name']
    ag_folder = USIT_folder.get_items(limit=100, offset=0)[0]

    for item in ag_folder.get_items(limit=100, offset=0):
        try:
            file_name = item.get()['name']
            download_url = item.get_shared_link_download_url()
            code = download_url.split("/static/")[1].split(".")[0]

            file_url = "https://app.box.com/embed/s/" + code
            # print(file_url)

            file = AnalystFile(name = file_name, filePath = file_url, owner = 'Box')
            db.session.add(file)
            db.session.commit()
            print("I added " + file_name + " to the database!")
        except:
            print("Cannot add file!")

    # print(client.search('USIT Platform', limit=100, offset=0)[0].metadata())

    # shared_folder = root_folder.create_subfolder('shared_folder')

    # uploaded_file = shared_folder.upload('/path/to/file')
    # shared_link = shared_folder.get_shared_link()

@manager.command
def get_fund_materials():
    #authentication
    #link to get Box developer token: https://app.box.com/developers/console/app/524950/configuration
    dev_token = str(raw_input("Developer token: "))

    oauth = OAuth2(
        client_id='l91tu18y9v1t9yth4w3xub87jbyln18k',
        client_secret='BYSJR7wr4Tl7Socbw3l87FYK01OOE91r',
        access_token=dev_token,
    )

    #go to fund folder
    client = Client(oauth)
    root_folder = client.folder(folder_id='0')
    USIT_folder = root_folder.get_items(limit=100, offset=0)[0]
    fund_folder = USIT_folder.get_items(limit=100, offset=0)[1]

    #adds files in each fund folder to the database
    for item in fund_folder.get_items(limit=100, offset=0):
        folder_name = item.get()['name']
        #print folder_name

        for file in item.get_items(limit=100, offset=0):
            try:
                download_url = file.get_shared_link_download_url()
                file_name = file.get()['name']
                code = download_url.split("/static/")[1].split(".")[0]
                file_url = "https://app.box.com/embed/s/" + code
                print(file_url)
                file = FundFile(name = file_name, filePath=file_url, owner='Box', fund= folder_name)
                db.session.add(file)
                db.session.commit()
                print("I added " + file_name + " to the database!")
            except:
                print("Cannot add file!")

@manager.command
def dropdb():
    if prompt_bool(
        "Are you sure you want to lose all your data"):
        db.drop_all()
        print('Dropped the database')

@manager.command
def get_attendance():
    members = User.query.filter_by(atLatestMeeting = True).all()
    count = 0

    for member in members:
        count += 1

    print("Numer of members: " + str(count))

@manager.command
def print_zero_dues():
    for user in User.query.all():
        if user.dues == 0:
            print("Name: " + user.firstName + " " + user.lastName)

@manager.command
def reset_attendance():
    email = str(input("Email: "))
    member = User.get_by_email(email)
    if not member is None:
        meetings_attended = int(input("How many meetings do you want to reset it to? "))
        member.attendance = meetings_attended
        member.atLatestMeeting = False
        db.session.commit()

@manager.command
def add_dues():
    addDues = True

    while addDues:
        email = str(raw_input("Email: "))
        member = User.get_by_email(email)
        if not member is None:
            dues = int(raw_input("How much dues do you want to reset it to? "))
            member.dues = dues
            db.session.commit()

        answer = raw_input("Do you want to continue (y/n)? ")
        if answer == 'y':
            addDues = True
        else:
            addDues = False

@manager.command
def get_dues():
    name_column_index = 4
    dues_column_index = 2

    trans_start_index = int(input("What row do you want to start at? "))
    trans_end_index = int(input("What row do you want to end at? "))

    transactions_unmatched = []

    for trans_index in range(trans_start_index, trans_end_index):
        fullName = str(transactions.cell(trans_index, name_column_index).value).lower()
        foundMatch = False

        for member in User.query.all():
            fullNameRecord = (member.firstName + " " + member.lastName).lower()
            if fullNameRecord == fullName:
                foundMatch = True
                if member.dues == 0:
                    dues = transactions.cell(trans_index, dues_column_index).value
                    member.dues = int(dues.split("$")[1].split(".")[0])
                    db.session.commit()
                    print("I have added", dues, "to", fullName)

        if not foundMatch:
            transactions_unmatched.append(fullName)


    print("I have not found matches for the following names: ")
    for name in transactions_unmatched:
        print(name)

@manager.command
def get_from_sheet():
    eid_column_index = 3
    firstname_column_index = 1
    lastname_column_index = 2
    email_column_index = 4
    attendance_column_index = 8

    row_index = int(input("What is the starting row? "))
    end_index = int(input("What is the ending row? "))

    for index in range(row_index, end_index):
        if str(sheet.cell(index, eid_column_index).value):
            eid = str(sheet.cell(index, eid_column_index).value).lower()
            member = User.get_by_eid(eid)
            if not member is None:
                member.firstName = str(sheet.cell(index, firstname_column_index).value).title()
                member.lastName = str(sheet.cell(index, lastname_column_index).value).title()
                member.email = str(sheet.cell(index, email_column_index).value).lower()
                # member.attendance = int(sheet.cell(index, 11).value)
                # if sheet.cell(index, 8).value:
                    # member.dues = int(sheet.cell(index, 8).value)
                # else:
                    # member.dues = 0
                # member.dues = 0
                if sheet.cell(index, attendance_column_index).value:
                    member.attendance = int(sheet.cell(index, attendance_column_index).value)
                else:
                    member.attendance = 0
                # if sheet.cell(index, 5).value:
                    # member.year = str(sheet.cell(index, 5).value)
                # else:
                    # member.year = ""
                # if sheet.cell(index, 7).value:
                    # member.comments = str(sheet.cell(index, 7).value)
                # else:
                    # member.comments = ""
                member.rowOnSheet = index
                db.session.commit()
                print("I updated the information for", member.firstName, member.lastName)
            else:
                eid = str(sheet.cell(index, eid_column_index).value).lower()
                firstName = str(sheet.cell(index, firstname_column_index).value).title()
                lastName = str(sheet.cell(index, lastname_column_index).value).title()
                email = str(sheet.cell(index, email_column_index).value).lower()
                dues = 0
                # # attendance = int(sheet.cell(index, 11).value)
                # if sheet.cell(index, 8).value:
                #   dues = int(sheet.cell(index, 8).value)
                # else:
                #   dues = 0
                if sheet.cell(index, attendance_column_index).value:
                    attendance = int(sheet.cell(index, attendance_column_index).value)
                else:
                    attendance = 0
                # if sheet.cell(index, 5).value:
                #   year = str(sheet.cell(index, 5).value)
                # else:
                #   year = ""
                # if sheet.cell(index, 7).value:
                #   comments = str(sheet.cell(index, 7).value)
                # else:
                #   comments = ""
                rowOnSheet = index
                member = Member(eid = eid, firstName = firstName, lastName = lastName, email = email, dues = dues, attendance = attendance, rowOnSheet = rowOnSheet)
                db.session.add(member)
                print("I added", firstName, lastName)
                db.session.commit()
                # print("He/she is not in the database.")
        else:
            print("There is no one on row", index)

@manager.command
def write_to_sheet():
    row_index = 2
    end_index = 1000

    new_row_start = int(input("What row do you want to start at? "))

    column = int(input("What is the column for attendance? "))

    for member in User.query.filter_by(atLatestMeeting = True):
        if member.rowOnSheet is not None and member.rowOnSheet != 0:
            print("Row: ", member.rowOnSheet)
            sheet.update_cell(member.rowOnSheet, column, "X")
            member.atLatestMeeting = False
            if not str(sheet.cell(member.rowOnSheet, 1).value):
                sheet.update_cell(member.rowOnSheet, 1, member.firstName)
            if not str(sheet.cell(member.rowOnSheet, 2).value):
                sheet.update_cell(member.rowOnSheet, 2, member.lastName)
            if not str(sheet.cell(member.rowOnSheet, 4).value):
                sheet.update_cell(member.rowOnSheet, 4, member.email)
            print("I updated attendance in column", column, "for", member.firstName, member.lastName)
            db.session.commit()
        else:
            for index in range(new_row_start, end_index):
                if not str(sheet.cell(index, 1).value) and not str(sheet.cell(index, 2).value) and not str(sheet.cell(index, 3).value):
                    print("Row: ", index)
                    sheet.update_cell(index, 1, member.firstName)
                    sheet.update_cell(index, 2, member.lastName)
                    sheet.update_cell(index, 3, member.eid)
                    sheet.update_cell(index, 4, member.email)
                    sheet.update_cell(index, column, "X")
                    member.rowOnSheet = index
                    member.atLatestMeeting = False
                    db.session.commit()
                    print("I added in the info for", member.firstName, member.lastName, "and updated attendance in column", column)
                    break;

@manager.command
def reset_vote_table():
    Vote.__table__.drop(db.session.bind)
    Vote.__table__.create(db.session.bind)

@manager.command   
def create_vote_table():
    Vote.__table__.create(db.session.bind)

@manager.command
def create_recent_vote_table():
    RecentVote.__table__.create(db.session.bind)

@manager.command
def reset_recent_vote_table():
    RecentVote.__table__.drop(db.session.bind)
    RecentVote.__table__.create(db.session.bind)

@manager.command
def count_no_names():
    count = 0
    for user in User.query.all():
        if not user.firstName or not user.lastName:
            count += 1
    print(str(count) + " people have no names in the database.")

@manager.command
def get_votes():
    ticker = str(raw_input("Ticker: ")).upper()

    numLong = Vote.query.filter_by(ticker=ticker, position="Yes - Long").count()
    numShort = Vote.query.filter_by(ticker=ticker, position="No - Short").count()
    numAbstain = Vote.query.filter_by(ticker=ticker, position="No - No Position").count()
    print("--------------- " + ticker + " Summary ---------------")
    print("Long: " + str(numLong) + ", Short: " + str(numShort) + ", Abstain: " + str(numAbstain))
    print("Number of Votes: " + str(numLong + numShort + numAbstain))

    for vote in Vote.query.filter_by(ticker=ticker):
        print("Email: " + vote.email + ", Vote: " + vote.position)

    print("--------------------------------------------")

@manager.command
def print_stocks():
    for stock in Stock.query.all():
        print(str(stock.id) + ". Stock: " + stock.ticker + ", " + str(stock.price))

    for ticker in Tickers.query.all():
        print(str(ticker.id) + ". Ticker: " + ticker.ticker + ", " + str(ticker.startingPrice) + ", " + str(ticker.short))

    for user in User.query.all():
        for stock in user.stocks:
            print(str(stock.id) + ". " + user.email + " voted for " + stock.ticker + ". Short? " + str(stock.short))
        for transaction in user.transactions:
            transTicker = Tickers.query.filter_by(ticker=transaction.ticker).first()
            if transTicker is not None:
                print(str(transaction.id) + ". " +  user.email + " took a position on " + transaction.ticker + " with return of " + str(transaction.returns/transTicker.startingPrice) + "%.")
            else:
                print(str(transaction.id) + ". " +  user.email + " took a position on " + transaction.ticker + " with return of $" + str(transaction.returns) + ".")

@manager.command
def print_returns():
    for user in User.query.all():
        print(str(user.email) + " has a return of " + str(user.ret) + "%.")

@manager.command
def delete_transactions():
    continue_or_not = 'y'

    while continue_or_not == 'y':
        id = int(input("What is the ID? "))
        transaction = Transactions.query.filter_by(id = id).first()
        db.session.delete(transaction)
        db.session.commit()
        continue_or_not = str(raw_input("Do you want to continue (y/n)? ")).lower()

@manager.command
def print_transaction_details():
    trans_id = int(input("What is the ID? "))
    transaction = Transactions.query.filter_by(id = trans_id).first()
    user = User.query.filter_by(id = transaction.user_id).first()
    print("User: " + user.email)
    print("Ticker: " + transaction.ticker)
    print("End price: " + str(transaction.end_price))
    print("Return: " + str(transaction.returns))

@manager.command
def refreshdb():
        # Refresh the stored information for each stock
    for stock in Tickers.query.all():
        create_stock_info(stock)
    # for stock in Transactions.query.all():
    #     create_stock_info(stock)
        
    # Refresh the score and ranks for each student
    for student in User.query.all():
        numStocks = update_ret(student, student.stocks, student.transactions)
        update_score(student, student.ret, numStocks)

@manager.command
def get_user_emails():
    for user in User.query.all():
        print(user.email)

@manager.command
def add_analyst_group():
    email_input = str(raw_input("What is the GM's email? "))
    student = User.query.filter(func.lower(User.email) == func.lower(email_input)).first()
    if student is not None:
        analyst_group = str(raw_input("Please list the senior analyst for {}: ".format(student.email)))
        student.analyst = analyst_group
        db.session.commit()

        student_new = User.query.filter_by(email = email_input).first()
        print(student.analyst + " is {}'s SA.".format(student.email))

@manager.command
def add_fund():
    email_input = str(raw_input("What is the GM's email? "))
    student = User.query.filter_by(email = email_input).first()
    if student is not None:
        fullName = student.firstName + " " + student.lastName
        fund = str(raw_input("Please list the fund for {}: ".format(fullName)))
        student.fund = fund
        db.session.commit()

    student_new = User.query.filter_by(email = email_input).first()
    fullName = student.firstName + " " + student.lastName
    print(student.fund + " is {}'s fund.".format(fullName))

@manager.command
def addstock():
    name = str(raw_input("What is the file name? "))
    symbol = str(raw_input("Ticker: "))
    price = int(raw_input("Starting price: $"))
    num_votes = int(raw_input("How many votes were there? "))

    wb = load_workbook(filename=name)
    ws = wb.active

    for index in range(1, num_votes + 1):
        if str(ws['B' + str(index)].value) == 'Long':
            if  Tickers.query.filter_by(short = False, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = False, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=False)
        elif str(ws['B' + str(index)].value) == 'Short':
            if Tickers.query.filter_by(short = True, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = True, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=True)
        if User.query.filter_by(email=str(ws['A'+str(index)].value).lower()) != None:
            student = User.query.filter_by(email=str(ws['A'+str(index)].value)).first()
            add_stock(student, stock)

@manager.command
def fixdb():
    num_students = User.query.count()
    num_students_wo_id = 1
    for student in User.query.all():
        if student.id == None:
            student.id == num_students + num_students_wo_id
            num_students_wo_id += 1
            print("I fixed", student.firstName, "!")

        db.session.commit()

@manager.command
def add_users_from_voting():
    users = get_users_from_vote()

    print(len(users))

    for user in users:
        try:
            db.session.add(User(id = user.id,
                                eid = '',
                                firstName = user.firstName,
                                lastName = user.lastName,
                                email = user.email,
                                attendance = 0,
                                dues = 0,
                                atLatestMeeting = False,
                                rowOnSheet = 0,
                                password_hash = user.password_hash,
                                ret = user.ret,
                                score = user.score,
                                active = False,
                                analyst = "",
                                fund = ""))

            db.session.commit()
            print("Added " + user.email)
        except:
            print("The session already contains this user.")
            print(str(user.id) + ". " + user.email)

@manager.command
def add_relationship_table():
    vote_ticker_identifier = get_ticker_identifier()
    vote_roles_users = get_roles_users()

    # for ticker_rel in vote_ticker_identifier:
    #     try:
    #         platformsession.execute(ticker_identifier.insert().values(user_id = int(ticker_rel.user_id), ticker_id = int(ticker_rel.ticker_id)))
    #         platformsession.commit()
    #         # db.session.commit()
    #         print("Added relationship: " + str(ticker_rel.user_id) + ", " + str(ticker_rel.ticker_id))
    #     except Exception as e:
    #         print(e)
    #         print("Doesn't work.")

    # for role_rel in vote_roles_users:
    #     try:
    #         platformsession.execute(roles_users.insert().values(user_id = int(role_rel.user_id), role_id = int(role_rel.role_id)))
    #         platformsession.commit()
    #         # db.session.commit()
    #         print("Added relationship: " + str(role_rel.user_id) + ", " + str(role_rel.role_id))
    #     except Exception as e:
    #         print(e)
    #         print("Doesn't work.")

# @manager.command
# def print_eids():
#     users = get_users_from_vote()

#     for user in users:
#         print(user.eid)

if __name__ == '__main__':
    manager.run()
