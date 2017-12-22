from __future__ import print_function # In python 2.7
import sys, os, operator
from datetime import datetime
from flask import Flask, render_template, request, flash, redirect, url_for, Markup
import forms
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_required, login_user, logout_user, current_user
from models import User, Tickers, Transactions, Role, Stock, Vote, RecentVote, FundFile, AnalystFile
from PlatformApp import db, app, login_manager, mail, cipher_suite, server
from yahoo_finance import Share
from pytz import timezone
from werkzeug import secure_filename
import requests
import time
import json
import smtplib
from itsdangerous import URLSafeTimedSerializer
from flask.ext.security import Security, SQLAlchemyUserDatastore, \
    UserMixin, RoleMixin, login_required
from openpyxl import load_workbook, Workbook
from sqlalchemy import func
from grampg import PasswordGenerator
import ntpath
import ssl
from lxml import html

try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    # Legacy Python that doesn't verify HTTPS certificates by default
    pass
else:
    # Handle target environment that doesn't support HTTPS verification
    ssl._create_default_https_context = _create_unverified_https_context


basedir = os.path.abspath(os.path.dirname(__file__))
date = time.strftime("%m/%d/%Y")
current_analysts = ['Srija Nalla']
current_funds = ['TMT']

# app = Flask(__name__)
# app.config['SECRET_KEY'] = '~t\x86\xc9\x1ew\x8bOcX\x85O\xb6\xa2\x11kL\xd1\xce\x7f\x14<y\x9e'
# app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'usit.db')
# db = SQLAlchemy(app)


# Configure Flask Security
# TODO dont know if we need this line, I don't have roles
user_datastore = SQLAlchemyUserDatastore(db, User, Role)
security = Security(app, user_datastore)

passwords = (PasswordGenerator().of().between(4, 5, 'letters')
                                     .at_least(2, 'numbers')
                                     .length(7)
                                     .beginning_with('letters')
                                     .done())
# def refreshdb(ticker):
#     # Refresh the stored information for each stock
#     for stock in db.session.query(Tickers).filter_by(ticker = ticker).all():
#         create_stock_info(stock)

def refreshdb():
    # Refresh the score and ranks for each student
    for student in User.query.all():
        numStocks = update_ret(student, student.stocks, student.transactions)
        update_score(student, student.ret, numStocks)

    # Refresh the stored information for each stock
    for stock in Tickers.query.all():
        create_stock_info(stock)
    for stock in Transactions.query.all():
        create_stock_info(stock)

def create_stock_info(stock):
    info = get_info(stock.ticker)

    if db.session.query(Stock).filter_by(ticker = stock.ticker).count() > 0:
        found_stock = db.session.query(Stock).filter_by(ticker = stock.ticker).first()
        found_stock.name = info['name']
        found_stock.price = info['price']
        found_stock.datetime = info['datetime']
        found_stock.change = info['gain']
        found_stock.percentChange = info['percentchange']
        print("I updated the info for: ", stock.ticker)
    else:
        new_stock = Stock(ticker = stock.ticker, name = info['name'], price = info['price'], datetime = info['datetime'], change = info['gain'], percentChange = info['percentchange'])
        db.session.add(new_stock)
        print("I added the info for: ", stock.ticker)

    db.session.commit()

def sort_users_by_score(item):
    return item.score

def get_json(ticker):
    url = "https://www.google.com/finance/info?q=NSE:{}".format(ticker)

    result = requests.get(url).text.split("// ")

    if len(result) > 1:
        rjson = json.loads(result[1])
    else:
        rjson = json.loads(result)

    return rjson

# def get_price(ticker):
#     rjson = get_json(ticker)
#     return float(rjson[0][u'l'])

def update_ret(self, stocks, transactions):
    ret = 0
    totalStocks = 0
    average_ret = 0

    for stock in stocks:
        price = float(get_price(stock.ticker))

        if (stock.short):
            ret += (stock.startingPrice - price)/stock.startingPrice
        else:
            ret += (price - stock.startingPrice)/stock.startingPrice

        totalStocks += 1

    for trans in transactions:
        # if trans.returns == 0:
        #     ret += 0
        # else:
        #     ret += trans.returns
        #     totalStocks += 1

        transTicker = Tickers.query.filter_by(ticker=trans.ticker).first()

        if not transTicker is None:
            if trans.returns == 0:
                ret += 0
            else:
                ret += trans.returns / transTicker.startingPrice
                totalStocks += 1
            # if(transTicker.short):
            #     ret += (transTicker.startingPrice - trans.end_price)/transTicker.startingPrice
            # else:
            #     ret += (trans.end_price - transTicker.startingPrice)/transTicker.startingPrice

            # totalStocks += 1
        else:
            ret += 0

    if totalStocks != 0:
        average_ret = ret/totalStocks
    else:
        average_ret = 0

    db.session.query(User).filter_by(id=self.id).first().ret = average_ret
    db.session.commit()

    return totalStocks

def update_score(self, ret, numStocks):
    if numStocks == 0:
        coins = 0
    else:
        lever = 20
        levered_returns = ret * lever
        coins = numStocks * (1 + levered_returns) 

    print("NUM COINS: ", coins)
    db.session.query(User).filter_by(id=self.id).first().score = coins
    db.session.commit()

def add_stock(self, stock):
    db.session.query(User).filter_by(id=self.id).first().stocks.append(stock)
    db.session.commit()

@login_manager.user_loader
def load_user(userid):
    return User.query.get(int(userid))

@app.route('/', methods=['GET', 'POST'])
#@app.route('/login', methods=['GET', 'POST'])
def index():
    form = forms.LoginForm()
    setForm = forms.SignUpForm()
    if request.method=='POST' and request.form['btn'] == 'log in':
        email = form.eid.data
        password = cipher_suite.encrypt(form.password.data.encode())
        user = User.get_by_eid(email)
        if user is not None and user.check_password(password):
            login_user(user, False)
            return redirect(url_for('dashboard'))
        else:
            flash("Incorrect Email or Password")
            #return redirect(url_for('index'))
    elif request.method=='POST' and request.form['btn'] == 'Sign Up':
        print("IN SIGNUP")
        if validateSignUp():
            print("Creating User")
            user = User(email = setForm.setEmail.data,
                        eid = setForm.setEID.data,
                        firstName = setForm.firstName.data,
                        lastName = setForm.lastName.data,
                        password = setForm.setPassword.data,
                        ret = 0,
                        score = 0,
                        dues = 0,
                        atLatestMeeting = False,
                        rowOnSheet = 0, 
                        attendance = 0,
                        analyst = None)
            db.session.add(user)
            db.session.commit()
            return redirect(url_for('index'))
        return render_template('index.html', form=form, setForm=setForm, login="login-hide", signup="signup-show", login_active="", signup_active="active")
    return render_template('index.html', form=form, setForm=setForm, login="login", signup="signup", login_active="active", signup_active="")

def validateSignUp():
    setForm = forms.SignUpForm()
    ok = True
    if User.query.filter_by(email=setForm.setEmail.data).first():
        flash("User with email already exists")
        ok = False
    if User.query.filter_by(eid=setForm.setEID.data).first():
        flash("User with EID already exists")
        ok = False
    if setForm.setPassword.data != setForm.setPassword2.data:
        flash("Passwords do NOT match")
        ok = False
    if (setForm.setEmail.data)[-10:] != "utexas.edu":
        flash("Must use utexas.edu email")
        ok = False
    if len(setForm.setPassword.data) < 6:
        flash("Password must be atleast six characters long")
        ok = False
    if not setForm.setEmail.data or not setForm.firstName.data or not setForm.lastName.data or not setForm.setPassword.data:
        flash("All fields are required")
        ok = False
    return ok

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/loading')
def loading():
    return render_template('loading.html')

#TODO need to calculate and send todays, total gains
@app.route('/dashboard')
@login_required
def dashboard():
    """Getting Position on Leaderboard"""
    allUsers = User.query.all()
    scores = {}
    emails = []
    for student in allUsers:
        scores[student.id] = student.score
        emails.append(student.email)
    "Sorts the dictionary by returns"
    score_tups = sorted(scores.items(), key=operator.itemgetter(1), reverse=True)
    users_tups = sorted(allUsers, key=sort_users_by_score, reverse=True)
    "Finds leadboard position"
    standing = 1
    for userid, score in score_tups:
        if userid is current_user.id:
            break
        standing += 1
    """Finished getting position"""
    prices = {}
    names = {}
    dates = {}
    changes = {}
    percentChanges = {}
    trends = {}
    totalGains = {}
    totalPercents = {}
    shorts = {}
    for stock in current_user.stocks:
        info = get_info_server(stock.ticker)
        prices[stock.ticker] = info['price']
        names[stock.ticker] = info['name']
        dates[stock.ticker] = info['datetime']
        #TODO make need to take into shorts in get_gain and get_percent_change
        changes[stock.ticker] = info['gain']
        percentChanges[stock.ticker] = info['percentchange']
        if(stock.short):
            totalGains[stock.ticker] = float('%.2f' % (stock.startingPrice - prices[stock.ticker]))
            if stock.startingPrice != 0:
                totalPercents[stock.ticker] = round(((stock.startingPrice - prices[stock.ticker])/stock.startingPrice)*100, 2)
            else:
                totalPercents[stock.ticker] = 0
            shorts[stock.ticker] = True
            changes[stock.ticker] *= -1
            percentChanges[stock.ticker] *= -1
        else:
            totalGains[stock.ticker] = float('%.2f' % (prices[stock.ticker] - stock.startingPrice))
            if stock.startingPrice != 0:
                totalPercents[stock.ticker] = round(((prices[stock.ticker] - stock.startingPrice)/stock.startingPrice)*100, 2)
            else:
                totalPercents[stock.ticker] = 0
            shorts[stock.ticker] = False
    startingPrices = {}
    for stock in Tickers.query.all():
        startingPrices[stock.ticker] = stock.startingPrice
    exitedStocks = Transactions.query.filter_by(user_id=current_user.id)
    exitedStocksNames = {}
    exitedStockDates = {}
    exitedGains = {}
    #TODO Have to calculate totalGains and totalPercents for exited stocks
    for stock in exitedStocks:
        info = get_info_server(stock.ticker)
        exitedStocksNames[stock.ticker] = info['name']
        exitedStockDates[stock.ticker] = info['datetime']
        exitedGains[stock.ticker] = stock.returns

    if not current_user.firstName or not current_user.lastName:
        needNames = True
    else:
        needNames = False

    if RecentVote.query.first() != None:
        if not RecentVote.query.first().close and Vote.query.filter_by(email=current_user.email,ticker=RecentVote.query.first().ticker).count() == 0:
            hasNotVoted = True
        else:
            hasNotVoted = False
    else:
        hasNotVoted = False

    if RecentVote.query.first() != None:
        numLong = Vote.query.filter_by(position="Yes - Long",ticker=RecentVote.query.first().ticker).count()
        numShort = Vote.query.filter_by(position="No - Short",ticker=RecentVote.query.first().ticker).count()
        numAbstain = Vote.query.filter_by(position="No - No Position",ticker=RecentVote.query.first().ticker).count()
        labels = ['Yes - Long', 'No - short', 'No - No Position']
        values = [numLong, numShort, numAbstain]
        colors = ["#1A2930", "#F7CE3E", "#FFFFFF"]
        displayVotes = True
    else:
        numLong = 0
        numShort = 0
        numAbstain = 0
        labels = []
        values = []
        colors = []
        displayVotes = False

    if not ((numLong + numShort + numAbstain) > 0):
        displayVotes = False

    if RecentVote.query.first():
        if RecentVote.query.first() != None:
            outstandingTicker = RecentVote.query.first().ticker
        else:
            outstandingTicker = ''

        if RecentVote.query.first().close == False:
            closeVotingButton = True
        else:
            closeVotingButton = False
    else:
        outstandingTicker = ''
        closeVotingButton = False

    if not(current_user.analyst is None) and current_user.analyst in current_analysts:
        showAnalystGroupButton = True
    else:
        showAnalystGroupButton = False

    if not(current_user.fund is None) and current_user.fund in current_funds:
        showFundButton = True
    else:
        showFundButton = False

    if len(current_user.roles) > 0 and (current_user.roles[0].name == 'Officer' or current_user.roles[0].name=='Admin'):
        if current_user.roles[0].name == 'Admin':
            return render_template('dashboard.html', email = current_user.email, needNames = needNames, emails = emails, isOfficer = False, isAdmin=True, rankings = users_tups, stocks=current_user.stocks, prices=prices,
                names = names, score = current_user.score, totalReturn=current_user.ret, standing=standing, startingPrices=startingPrices,
                exitedStocks=exitedStocks, exitedStocksNames = exitedStocksNames, numExitedStocks = len(exitedStocksNames),
                numActiveStocks = len(current_user.stocks), firstName = current_user.firstName,
                lastName = current_user.lastName, dates = dates, exitedStockDates = exitedStockDates,
                changes = changes, percentChanges = percentChanges, totalGains = totalGains, totalPercents = totalPercents,
                exitedGains = exitedGains, outstandingVote = hasNotVoted, outstandingTicker = outstandingTicker, values = values, labels = labels, colors = colors, displayVotes = displayVotes, closeVotingButton = closeVotingButton, showAttendance = False, showAnalystGroup = False,
                showAnalystGroupButton = showAnalystGroupButton, showFund = False, showFundButton = showFundButton)
        else:
            return render_template('dashboard.html', email = current_user.email, needNames = needNames, emails = emails, isOfficer = True, isAdmin=False, rankings = users_tups, stocks=current_user.stocks, prices=prices,
                names = names, score = current_user.score, totalReturn=current_user.ret, standing=standing, startingPrices=startingPrices,
                exitedStocks=exitedStocks, exitedStocksNames = exitedStocksNames, numExitedStocks = len(exitedStocksNames),
                numActiveStocks = len(current_user.stocks), firstName = current_user.firstName,
                lastName = current_user.lastName, dates = dates, exitedStockDates = exitedStockDates,
                changes = changes, percentChanges = percentChanges, totalGains = totalGains, totalPercents = totalPercents,
                exitedGains = exitedGains, outstandingVote = hasNotVoted, outstandingTicker = outstandingTicker, values = values, labels = labels, colors = colors, displayVotes = displayVotes, closeVotingButton = closeVotingButton, showAttendance = False, showAnalystGroup = False,
                showAnalystGroupButton = showAnalystGroupButton, showFund = False, showFundButton = showFundButton)
    else:
        return render_template('dashboard.html', email = current_user.email, needNames = needNames, isOfficer=False, isAdmin=False, stocks=current_user.stocks, prices=prices,
            names = names, score = current_user.score, totalReturn=current_user.ret, standing=standing, startingPrices=startingPrices,
            exitedStocks=exitedStocks, exitedStocksNames = exitedStocksNames, numExitedStocks = len(exitedStocksNames),
            numActiveStocks = len(current_user.stocks), firstName = current_user.firstName,
            lastName = current_user.lastName, dates = dates, exitedStockDates = exitedStockDates,
            changes = changes, percentChanges = percentChanges, totalGains = totalGains, totalPercents = totalPercents,
            exitedGains = exitedGains, outstandingVote = hasNotVoted, outstandingTicker = outstandingTicker, values = values, labels = labels, colors = colors, displayVotes = displayVotes, closeVotingButton = closeVotingButton, showAttendance = False, showAnalystGroup = False,
            showAnalystGroupButton = showAnalystGroupButton, showFund = False, showFundButton = showFundButton)

def get_json(ticker):
    url = "https://www.google.com/finance/info?q=NSE:{}".format(ticker)

    result = requests.get(url).text.split("// ")

    if len(result) > 1:
        rjson = json.loads(result[1])
    else:
        rjson = json.loads(result)

    return rjson

def get_info_server(ticker):

    info = {}

    if db.session.query(Stock).filter_by(ticker = ticker).count() > 0:
        found_stock = db.session.query(Stock).filter_by(ticker = ticker).first()
        info['name'] = found_stock.name
        info['price'] = found_stock.price
        info['datetime'] = found_stock.datetime
        info['gain'] = found_stock.change
        info['percentchange'] = found_stock.percentChange
    else:
        info['name'] = "N/A"
        info['price'] = 0.00
        info['datetime'] = "N/A"
        info['gain'] = 0.00
        info['percentchange'] = 0.00

    return info    

def get_info(ticker):
    # rjson = get_json(ticker)
    info = {}

    # try:
    urlStock = "http://www.nasdaq.com/symbol/{}".format(ticker)
    page = requests.get(urlStock)
    tree = html.fromstring(page.content)

    # print(tree.xpath('//div[@id="qwidget_lastsale"]/text()'))

    price = round(float(tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0].split("$")[1]), 2)
    change = round(float(tree.xpath('//div[@id="qwidget_netchange"]/text()')[0]), 2)
    pchange = round(float(tree.xpath('//div[@id="qwidget_percent"]/text()')[0].split("%")[0]), 2)
    change_class = tree.xpath('//div[@id="qwidget_netchange"]')[0].get('class')

    # price = get_info_server(ticker)['price']
    # change = get_info_server(ticker)['gain']
    # pchange = get_info_server(ticker)['percentchange']
    # change_class = 'qwidget-cents qwidget-Green'

    print(str(price), str(change), str(pchange))

    negative = True

    if change_class == 'qwidget-cents qwidget-Red':
        negative = True
    elif change_class == 'qwidget-cents qwidget-Green':
        negative = False
    else:
        print("I guess it does not exist...")
    
    # stock = Share(ticker)

    # 0: Get name
    url = "http://d.yimg.com/autoc.finance.yahoo.com/autoc?query={}&region=1&lang=en".format(ticker)
    print("Ticker: ", ticker)
    result = requests.get(url).json()

    for x in result['ResultSet']['Result']:
        if x['symbol'] == ticker:
            info['name'] = truncate(x['name'])

    if info['name'] == "" or info['name'] == None:
        raise ValueError('Did not obtain a real value!')


    # 1: Get price
    # info['price'] = float(rjson[0][u'l'])
    # info['price'] = get_price(ticker)
    info['price'] = price

    if info['price'] == 0 or info['price'] == None:
        raise ValueError('Did not obtain a real value!')

    # 2: Get datetime
    # info['datetime'] = rjson[0][u'lt']
    info['datetime'] = getdatetime(ticker)

    if info['datetime'] == "" or info['datetime'] == None:
        raise ValueError('Did not obtain a real value!')

    # 3: Get gain
    # change = rjson[0][u'c']
    # if change is None:
    #     info['gain'] = 0
    # c = change.split("+")
    # if (len(c) > 1):
    #     info['gain'] = float(c[1])
    # info['gain'] = float(change)

    if negative:
        change = -1 * change
    
    if change is None:
        info['gain'] = 0
    else:
        info['gain'] = change

    if info['gain'] == None:
        raise ValueError('Did not obtain a real value!')

        # 4: Get percent change
        # info['percentchange'] = float(rjson[0][u'cp'])
        # try:
        #     percentChange = stock.get_percent_change()
        #     percentChange = percentChange.split("%")[0]
        #     if len(percentChange.split("+")) > 1:
        #         percentChange = percentChange.split("+")[1]
        #     elif len(percentChange.split("-")) > 1:
        #         percentChange = percentChange.split("-")[1]

        #     info['percentchange'] = float(percentChange)
        # except:
        #     info['percentchange'] = stock.get_percent_change()
    
    if negative:
        info['percentchange'] = -pchange
    else:
        info['percentchange'] = pchange

    if info['percentchange'] == None:
        raise ValueError('Did not obtain a real value!')

    # except:
    #     if db.session.query(Stock).filter_by(ticker = ticker).count() > 0:
    #         found_stock = db.session.query(Stock).filter_by(ticker = ticker).first()
    #         info['name'] = found_stock.name
    #         info['price'] = found_stock.price
    #         info['datetime'] = found_stock.datetime
    #         info['gain'] = found_stock.change
    #         info['percentchange'] = found_stock.percentChange
    #     else:
    #         info['name'] = "N/A"
    #         info['price'] = 0.00
    #         info['datetime'] = "N/A"
    #         info['gain'] = 0.00
    #         info['percentchange'] = 0.00

    return info

def getdatetime(ticker):
    # fromAPIString = Share(ticker).get_trade_datetime()
    # parsedString = fromAPIString.split("+")[0]
    # fromAPIDate = datetime.strptime(parsedString, "%Y-%m-%d %H:%M:%S %Z")
    fromAPIDate = datetime.now()
    etzDate = timezone("US/Central").localize(fromAPIDate)
    returnDate = etzDate.strftime("%m/%d/%Y %I:%M %p %Z")
    return returnDate

def get_price(ticker):
    urlStock = "http://www.nasdaq.com/symbol/{}".format(ticker)
    page = requests.get(urlStock)
    tree = html.fromstring(page.content)

    #print("Price: " + tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0])
    if len(tree.xpath('//div[@id="qwidget_lastsale"]/text()')) > 0: 
        if len(tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0].split("$")) > 1:
            price = round(float(tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0].split("$")[1]), 2)
        else:
            price = round(float(tree.xpath('//div[@id="qwidget_lastsale"]/text()')[0]), 2)
    else:
        price = 0
    return price

def truncate(name):
    if (len(name) > 20):
        words = name.split(" ")
        length = 0
        return_name = ""
        for x in words:
            if (len(x) + length + 1 < 20):
                return_name += x + " "
                length += len(x) + 1
            else:
                break
        return return_name

    return name

#TODO take into account shorts in color of exited positions
@app.route('/exitPosition/<int:exitIndex>')
def exitPosition(exitIndex):
    #print(current_user.id)  #user_id
    #print(current_user.stocks[exitIndex-1].id)  #ticker_id
    """Deletes position from Active Positions"""
    user = User.query.filter_by(id=current_user.id).first()
    exitPosition = current_user.stocks[exitIndex-1]
    print()
    print()
    print("EXIT INDEX: ", exitIndex)
    print()
    print()
    current_user.stocks.pop(exitIndex-1)
    db.session.commit()

    """Need to add to transactions table"""
    t = datetime.now()
    today = str(t.month) + "/" + str(t.day) + "/" + str(t.year)
    exitPrice = float(get_info(exitPosition.ticker)['price'])

    if (exitPosition.short):
        exitReturn = exitPosition.startingPrice - exitPrice
    else:
        exitReturn = exitPrice - exitPosition.startingPrice

    transaction = Transactions(user_id=current_user.id,
                ticker=str(exitPosition.ticker),
                date=today,
                end_price = exitPrice,
                returns = exitReturn)
    db.session.add(transaction)
    db.session.commit()
    return redirect(url_for('dashboard'))


@app.route('/loggedin')
def loggedin():
    return render_template('loggedin.html')

@app.route('/upload')
def upload():
   return render_template('upload.html')

@app.route('/signin', methods=['GET', 'POST'])
@login_required
def log_in():
    if current_user.check_attendance(current_user.dues, current_user.attendance + 1):
        current_user.attendance += 1
        current_user.atLatestMeeting = True
        db.session.commit()

        return render_template('dashboard.html', showAttendance = True, showAnalystGroup = False, showFund = False, firstName = current_user.firstName, lastName = current_user.lastName, allowedIn = True, attendance = current_user.attendance, dues = current_user.dues, date = date)
    else:
        return render_template('dashboard.html', showAttendance = True, showAnalystGroup = False, showFund = False, firstName = current_user.firstName, lastName = current_user.lastName, allowedIn = False, attendance = current_user.attendance, dues = current_user.dues, date = date)

@app.route('/analystgroup', methods=['GET', 'POST'])
@login_required
def analyst_group():
    if not(current_user.analyst is None) and current_user.analyst in current_analysts:
        if len(current_user.roles) > 0 and (current_user.roles[0].name == 'Officer' or current_user.roles[0].name=='Admin'):
            addFiles = True
        else:
            addFiles = False
        analystFiles = AnalystFile.query.all()
        analystFileURLs = []
        for file in analystFiles:
            analystFileURLs.append(file.filePath)
        return render_template('dashboard.html', showAnalystGroup = True, showAttendance = False, showFund = False, firstName = current_user.firstName, lastName = current_user.lastName, analystFiles = analystFiles, URLs = analystFileURLs, showAGFileButton = addFiles)
    return redirect(url_for('dashboard'))

@app.route('/fund', methods=['GET', 'POST'])
@login_required
def fund():
    if not(current_user.fund is None) and current_user.fund in current_funds:
        return render_template('dashboard.html', showAnalystGroup = False, showAttendance = False, showFund = True, firstName = current_user.firstName, lastName = current_user.lastName)
    return redirect(url_for('dashboard'))

@app.route('/uploader', methods = ['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        f = request.files['file']
        try:
            f.save(secure_filename(f.filename))
        except:
            return 'No file was uploaded.'
        addstock(f.filename, request.form['ticker'], float(request.form['price']))
        return redirect('dashboard')

def addstock(name, symbol, price):
    wb = load_workbook(filename=name)
    ws = wb.active

    index = 1

    while ws['A' + str(index)].value:
        choice = str(ws['B' + str(index)].value)
        print("Looking at " + str(ws['A'+str(index)].value) + ".")
        if choice == 'Yes - long':
            if  Tickers.query.filter_by(short = False, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = False, ticker = symbol).first()
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=False)
                # refreshdb(symbol)
        elif choice == 'No - short':
            if Tickers.query.filter_by(short = True, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = True, ticker = symbol).first()
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=True)
                # refreshdb(symbol)
        elif choice == 'No - no position':
            if User.query.filter(func.lower(User.email) == func.lower(str(ws['A'+str(index)].value))).first() != None:
                student = User.query.filter(func.lower(User.email) == func.lower(str(ws['A'+str(index)].value))).first()
                if Transactions.query.filter_by(user_id = student.id, ticker = symbol).first() == None:
                    t = datetime.now()
                    today = str(t.month) + "/" + str(t.day) + "/" + str(t.year)
                    transaction = Transactions(user_id=student.id,
                                               ticker=symbol,
                                               date=today,
                                               end_price = float(get_price(symbol)),
                                               returns = 0)
                    db.session.add(transaction)
                    db.session.commit()

        if choice == 'Yes - long' or choice == 'No - short':
            if User.query.filter(func.lower(User.email) == func.lower(str(ws['A'+str(index)].value))).first() != None:
                student = User.query.filter(func.lower(User.email) == func.lower(str(ws['A'+str(index)].value))).first()
                add_stock(student, stock)
            else:
                member = Role(name="Member", description="Just a general user to start off with")
                password = passwords.generate()
                db.session.add(User(firstName="", lastName="",email=str(ws['A'+str(index)].value), password=password, roles=[member]))
                db.session.commit()
                msg = "Hello,\n\nThank you for being an active participant at USIT's general meetings and voting in the Voting Challenge! Since you do not already have an account, we have created an account for you! We highly recommend that you visit vote.texasusit.org/reset to change your password to something more familiar. Below, we have inserted your new temporary password.\n\nEmail: {}\nPassword: {}\n\nThank you,\nUSIT Team".format(str(ws['A'+str(index)].value), password)
                server.sendmail('votingchallenge@usiteam.org', str(ws['A'+str(index)].value).lower(), msg)
                student = User.query.filter(func.lower(User.email) == func.lower(str(ws['A'+str(index)].value))).first()
                add_stock(student, stock)

        index += 1

    refreshdb()

@app.route('/newagmaterial', methods = ['POST', 'GET'])
def newagmaterial():
    if request.method == 'POST':
        agmaterial = AnalystFile(name = request.form['name'], filePath = request.form['filepath'], owner = current_user.email)
        db.session.add(agmaterial)
        db.session.commit()

        return redirect('analystgroup')

@app.route('/newvote', methods = ['POST', 'GET'])
def newvote():
    if request.method == 'POST':
        RecentVote.query.delete()
        recentVote = RecentVote(ticker=request.form['ticker'], close = False)
        db.session.add(recentVote)
        db.session.commit()

        return redirect('dashboard')

@app.route('/vote', methods = ['POST', 'GET'])
def vote():
    if request.method == 'POST':
        addstockvote(request.form['email'], request.form['votechoice'])
        return redirect('dashboard')

def addstockvote(email, choice):
    symbol = RecentVote.query.first().ticker
    price = get_price(symbol)

    vote = Vote(ticker = symbol, email = email, position = choice)
    db.session.add(vote)
    db.session.commit()

    if choice == 'Yes - Long':
        if  Tickers.query.filter_by(short = False, ticker = symbol).count() > 0:
            stock = Tickers.query.filter_by(short = False, ticker = symbol).first()
        else:
            stock = Tickers(ticker=symbol, startingPrice=price, short=False)
            # refreshdb(symbol)
    elif choice == 'No - Short':
        if Tickers.query.filter_by(short = True, ticker = symbol).count() > 0:
            stock = Tickers.query.filter_by(short = True, ticker = symbol).first()
        else:
            stock = Tickers(ticker=symbol, startingPrice=price, short=True)
            # refreshdb(symbol)
    elif choice == 'No - No Position':
        if User.query.filter(func.lower(User.email) == func.lower(email)).first() != None:
            student = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            if Transactions.query.filter_by(user_id = student.id, ticker = symbol).first() == None:
                t = datetime.now()
                today = str(t.month) + "/" + str(t.day) + "/" + str(t.year)
                transaction = Transactions(user_id=student.id,
                                           ticker=symbol,
                                           date=today,
                                           end_price = float(get_price(symbol)),
                                           returns = 0)
                db.session.add(transaction)
                db.session.commit()

    if choice == 'Yes - Long' or choice == 'No - Short':
        if User.query.filter(func.lower(User.email) == func.lower(email)).first() != None:
            student = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            add_stock(student, stock)
        else:
            member = Role(name="Member", description="Just a general user to start off with")
            password = passwords.generate()
            db.session.add(User(firstName="", lastName="",email=str(email), password=password, roles=[member]))
            db.session.commit()
            msg = "Hello,\n\nThank you for being an active participant at USIT's general meetings and voting in the Voting Challenge! Since you do not already have an account, we have created an account for you! We highly recommend that you visit vote.texasusit.org/reset to change your password to something more familiar. Below, we have inserted your new temporary password.\n\nEmail: {}\nPassword: {}\n\nThank you,\nUSIT Team".format(email, password)
            server.sendmail('votingchallenge@usiteam.org', str(email).lower(), msg)
            student = User.query.filter(func.lower(User.email) == func.lower(email)).first()
            add_stock(student, stock)

@app.route('/closevoting', methods=['POST', 'GET'])
def closevoting():
    if request.method == 'POST':
        RecentVote.query.first().close = True
        db.session.commit()
        return redirect('dashboard')

@app.route('/role', methods = ['POST'])
def change_role():
    if request.method == 'POST':
        email = str(request.form['email'])
        role = str(request.form['newrole']).title()
        new_role(email, role)
        return redirect('dashboard')

def new_role(address, newrole):
    if User.query.filter_by(email=address).first() != None:
        student = User.query.filter_by(email=address).first()
        if newrole == 'Delete User':
            db.session.delete(student)
            db.session.commit()
            return
        if Role.query.filter_by(name=newrole).first() != None:
            role = Role.query.filter_by(name=newrole).first()
        else:
            role = Role(name=newrole, description="")

        if len(student.roles) > 0:
            student.roles[0] = role
        else:
            student.roles.append(role)
            
        db.session.commit()    

@app.route('/name', methods=['POST'])
def change_name():
    if request.method == 'POST':
        email = str(request.form['email'])
        firstName = str(request.form['firstName']).title()
        lastName = str(request.form['lastName']).title()
        change_name(email, firstName, lastName)
        return redirect('dashboard')

def change_name(email, firstName, lastName):
    if User.query.filter_by(email = email).first() != None:
        student = User.query.filter_by(email=email).first()
        student.firstName = firstName
        student.lastName = lastName
        db.session.commit()
    else:
        print("Student was not found for some reason.")

if __name__ == '__main__':
    app.run(debug=True)
